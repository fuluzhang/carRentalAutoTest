import openpyxl,requests

class InterfaceTest:

    def test(self,path,sheetname):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheetname]
        data = [sheet.cell(i, 8).value for i in range(2, 29)]
        url = [sheet.cell(i, 6).value for i in range(2, 29)]
        expect = [sheet.cell(i, 10).value for i in range(2,29)]
        result = [sheet.cell(i, 11) for i in range(2, 29)]
        print(result[0])
        print(sheet['K2'])
        try:
            for i in range(27):
                r = requests.post(url[i], data=data[i])
                # result[i] = r.text
                # sheet['K%d+1'%i] = r.text
                sheet['K2'] ='4555555555555'
                assert r.text == expect[i]
        except Exception as e:
            print(e)
        finally:
            workbook.save(path)

if __name__ == '__main__':
    InterfaceTest().test(r'D:\ApiAutoTest\LearnRequests\金融接口测试用例_1.1(2).xlsx', 'Sheet1')

